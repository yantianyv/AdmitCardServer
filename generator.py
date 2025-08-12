import argparse
import json
import os
import sys
import csv
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus.flowables import Flowable, HRFlowable

# 默认配置模板
default_config = {
    "exam_name": "默认考试",
    "settings": {"render_photo_frame": False, "auto_extend": False},
    "exam_schedule": [{"subject": "科目", "time": "时间"}],
    "exam_notes": ["注意事项"],
}

# 全局缓存字体和样式
_font_registered = False
_styles_cache = None

def _init_global_styles():
    global _font_registered, _styles_cache
    if not _font_registered:
        pdfmetrics.registerFont(TTFont("WenQuanYi", "fonts/wqy-microhei.ttc"))
        _font_registered = True

    if _styles_cache is None:
        styles = getSampleStyleSheet()
        styles["Title"].fontName = "WenQuanYi"
        styles["Title"].fontSize = 24
        styles["Title"].textColor = colors.HexColor("#333333")
        styles["Heading1"].fontName = "WenQuanYi"
        styles["Heading2"].fontName = "WenQuanYi"
        styles["Normal"].fontName = "WenQuanYi"
        styles["Normal"].fontSize = 14
        styles.add(
            ParagraphStyle(
                name="Chinese",
                fontName="WenQuanYi",
                fontSize=14,
                leading=20,
                spaceBefore=6,
                spaceAfter=6,
            )
        )
        _styles_cache = styles
    return _styles_cache


class PhotoBox(Flowable):
    def __init__(self, width, height, text=""):
        Flowable.__init__(self)
        self.width = width
        self.height = height
        self.text = text

    def draw(self):
        self.canv.setDash(2, 2)
        self.canv.rect(0, 0, self.width, self.height)
        self.canv.setDash(1, 0)
        self.canv.setFont("WenQuanYi", 10)
        text_width = self.canv.stringWidth(self.text, "WenQuanYi", 10)
        self.canv.drawString(
            (self.width - text_width) / 2, self.height / 2 - 5, self.text
        )


def load_config(config_name="default"):
    config_path = os.path.join("config", f"{config_name}.json")

    global default_config

    if not os.path.exists(config_path):
        os.makedirs(os.path.dirname(config_path), exist_ok=True)
        with open(config_path, "w", encoding="utf-8") as f:
            json.dump(default_config, f, indent=2, ensure_ascii=False)
        print("\033[1;33m警告：配置文件不存在，已生成默认配置文件。\033[0m")
        return default_config

    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def validate_csv_structure(headers):
    if headers[0] != "姓名" or headers[1] != "身份证号":
        raise ValueError("CSV文件第一列应为'姓名'，第二列应为'身份证号'")


def read_csv_data(csv_path):
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        headers = next(reader)
        validate_csv_structure(headers)
        
        data = []
        for row in reader:
            student = {
                "name": row[0],
                "id": row[1],
                "fields": row[2:],
                "field_headers": headers[2:],
            }
            data.append(student)
    return data


def convert_excel_to_csv(excel_path):
    """将Excel文件转换为CSV，如果目标文件已存在则询问用户"""
    csv_path = os.path.splitext(excel_path)[0] + '.csv'
    
    # 检查CSV文件是否已存在
    if os.path.exists(csv_path):
        print(f"\n发现已存在的CSV文件: {csv_path}")
        while True:
            choice = input("是否覆盖现有文件？(y/n): ").strip().lower()
            if choice == 'y':
                break
            elif choice == 'n':
                new_name = input("请输入新的文件名(不带扩展名): ").strip()
                csv_path = os.path.join(os.path.dirname(excel_path), f"{new_name}.csv")
                if os.path.exists(csv_path):
                    print("该文件名也已存在，请重新选择")
                    continue
                break
            else:
                print("请输入y或n")
                continue

    wb = load_workbook(excel_path)
    ws = wb.active
    
    with open(csv_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
    
    print(f"已生成CSV文件: {csv_path}")
    return csv_path


def read_student_data(input_file):
    if input_file.lower().endswith(('.xls', '.xlsx')):
        print(f"检测到Excel文件 {input_file}，正在转换为CSV格式...")
        input_file = convert_excel_to_csv(input_file)
    
    return read_csv_data(input_file)


def generate_admit_card(student, config, output_dir="AdmitCards"):
    styles = _init_global_styles()
    filename = f"{student['id']}-{student['name']}.pdf"
    doc = SimpleDocTemplate(
        os.path.join(output_dir, filename),
        pagesize=A4,
        leftMargin=1 * cm,
        rightMargin=1 * cm,
        topMargin=1 * cm,
        bottomMargin=1 * cm,
    )

    elements = []

    if config.get("exam_name"):
        exam_name_style = ParagraphStyle(
            "ExamNameStyle",
            parent=styles["Title"],
            alignment=1,
            fontSize=18,
        )
        elements.append(Paragraph(f"{config['exam_name']}", exam_name_style))

    elements.append(Paragraph("准考证", styles["Title"]))

    info_elements = []
    if not student.get("name"):
        print(f"\033[1;33m警告：身份证号 {student['id']} 的姓名为空，已隐藏该字段\033[0m")
    else:
        info_elements.append(Paragraph(f"<b>姓名:</b> {student['name']}", styles["Normal"]))
        info_elements.append(Spacer(1, 0.5 * cm))

    if not student.get("id"):
        print(f"\033[1;33m警告：考生 {student['name']} 的身份证号为空，已隐藏该字段\033[0m")
    else:
        info_elements.append(Paragraph(f"<b>身份证号:</b> {student['id']}", styles["Normal"]))
        info_elements.append(Spacer(1, 0.5 * cm))

    for header, field in zip(student.get("field_headers", []), student["fields"]):
        if not field:
            print(f"\033[1;33m警告：考生 {student['name']} 的字段 '{header}' 为空，已隐藏该字段\033[0m")
        else:
            info_elements.append(Paragraph(f"<b>{header}:</b> {field}", styles["Normal"]))
            info_elements.append(Spacer(1, 0.5 * cm))

    render_photo = config["settings"]["render_photo_frame"]
    if render_photo:
        part1_table = Table(
            [[info_elements, PhotoBox(2.5 * cm, 3.5 * cm, "一寸照片粘贴处"), None]],
            colWidths=[10 * cm, 3 * cm, 1 * cm],
        )
    else:
        part1_table = Table([[info_elements, None]], colWidths=[13 * cm, 1 * cm])

    part1_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#4a86e8")),
                ("PADDING", (0, 0), (-1, -1), 16),
                ("ROUNDEDCORNERS", [4, 4, 4, 4]),
            ]
        )
    )

    elements.append(part1_table)
    elements.append(Spacer(1, 0.5 * cm))

    exam_info = [["科目", "考试时间"]]
    for schedule in config["exam_schedule"]:
        exam_info.append([schedule["subject"], schedule["time"]])
    exam_info.append(["", ""])

    row_heights = [1 * cm] * (len(exam_info) - 1) + [0.2 * cm]
    part2_table = Table(exam_info, colWidths=[4 * cm, 10 * cm], rowHeights=row_heights)
    part2_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "WenQuanYi"),
                ("FONTSIZE", (0, 0), (-1, -1), 14),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dddddd")),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f5f5f5")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#4a86e8")),
                ("PADDING", (0, 0), (-1, -1), 14),
                ("ROUNDEDCORNERS", [4, 4, 4, 4]),
                ("LINEBELOW", (-1, -1), (-1, -1), 1, colors.HexColor("#4a86e8")),
            ]
        )
    )

    elements.append(part2_table)
    elements.append(Spacer(1, 0.5 * cm))

    notes_elements = [Paragraph("注意事项：", styles["Heading2"])]
    for index, note in enumerate(config["exam_notes"], 1):
        note_style = ParagraphStyle(
            "NoteStyle",
            parent=styles["Normal"],
            leading=18,
            leftIndent=24,
            firstLineIndent=-12,
        )
        notes_elements.append(Paragraph(f"<b>{index}.</b> {note}", note_style))

    if config["settings"].get("auto_extend", True):
        used_height = (
            sum(e.wrap(doc.width, doc.height)[1] for e in elements) if elements else 0
        )
        remaining_height = max(0, (27.7 - 2) * cm - used_height)
        row_heights = [remaining_height]
    else:
        row_heights = None

    part3_table = Table([[notes_elements]], colWidths=[14 * cm], rowHeights=row_heights)
    part3_table.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#4a86e8")),
                ("PADDING", (0, 0), (-1, -1), 12),
                ("ROUNDEDCORNERS", [4, 4, 4, 4]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )

    elements.append(part3_table)

    doc.border = 0
    doc.borderPadding = 0
    doc.borderColor = None
    doc.borderStyle = None

    content_height = 0
    for e in elements:
        wrapped = e.wrap(doc.width, doc.height)
        content_height += wrapped[1] if wrapped else 0

    if content_height > doc.height:
        print(f"\033[1;33m警告：考生 {student['name']} (身份证号: {student['id']}) 的准考证内容过长（{content_height:.1f}点），可能无法在一页内完整显示。建议减少内容或调整格式。\033[0m")

    doc.build(elements)


def main():
    os.makedirs("config", exist_ok=True)
    os.makedirs("AdmitCards", exist_ok=True)

    if not os.path.exists(os.path.join("config", "default.json")):
        os.makedirs(os.path.dirname(os.path.join("config", "default.json")), exist_ok=True)
        with open(os.path.join("config", "default.json"), "w", encoding="utf-8") as f:
            json.dump(default_config, f, indent=2, ensure_ascii=False)
        print("\033[1;33m初次使用，已帮您生成默认配置文件，请修改后再次启动\033[0m")
        return default_config

    class CustomArgumentParser(argparse.ArgumentParser):
        def error(self, message):
            if "the following arguments are required" in message:
                print("请指定考生信息文件路径(CSV/XLS/XLSX)")
                self.exit(2)
            else:
                super().error(message)

    parser = CustomArgumentParser(description="准考证生成器", add_help=False)
    parser.add_argument("input_file", help="考生信息文件路径(CSV/XLS/XLSX)")
    parser.add_argument("-c", "--config", default="default", help="配置文件名（不带.json扩展名）")
    parser.add_argument("-h", "--help", action="help", help="显示帮助信息并退出")

    args = parser.parse_args()

    try:
        config = load_config(args.config)

        if not config.get('exam_name'):
            print("\033[1;33m警告：考试名称为空，将隐藏该字段\033[0m")

        students = read_student_data(args.input_file)
        total = len(students)
        print(f"开始生成{total}份准考证...")

        for i, student in enumerate(students, 1):
            generate_admit_card(student, config)
            progress = int(i / total * 100)
            progress_bar = "#" * (progress // 5) + "_" * (20 - progress // 5)
            sys.stdout.write(f"\r进度: {progress_bar} {progress}% ({i}/{total})")
            sys.stdout.flush()

        print(f"\n成功生成{total}份准考证到AdmitCards目录")
    except Exception as e:
        print(f"错误: {str(e)}")
        if isinstance(e, FileNotFoundError) and str(e).endswith(".json not found"):
            print("已自动创建默认配置文件，请按需修改后重新运行程序")


if __name__ == "__main__":
    main()

